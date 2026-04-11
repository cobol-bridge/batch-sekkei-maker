"""
Excelファイルの設計書生成
"""
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ヘッダー行の背景色（紺）
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="メイリオ", size=10)
BODY_FONT = Font(name="メイリオ", size=10)
SUBHEADER_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, name="メイリオ", size=10)

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _apply_header(cell, value: str):
    cell.value = value
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border


def _apply_body(cell, value):
    cell.value = value
    cell.font = BODY_FONT
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = thin_border


def generate_excel(parse_result, output_path: str):
    """
    パース結果からExcel設計書を生成する

    Args:
        parse_result: CobolParseResult
        output_path: 出力先ファイルパス
    """
    wb = openpyxl.Workbook()

    _create_cover_sheet(wb, parse_result)
    _create_io_sheet(wb, parse_result)
    _create_output_sheet(wb, parse_result)
    _create_layout_sheets(wb, parse_result)
    _create_flow_sheet(wb, parse_result)
    _create_exception_sheet(wb, parse_result)

    # デフォルトシートを削除
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(output_path)


def _create_cover_sheet(wb, parse_result):
    """シート1：表紙"""
    ws = wb.create_sheet("表紙", 0)
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 50

    title_font = Font(name="メイリオ", size=18, bold=True, color="1F4E79")
    ws["B2"].value = "バッチ処理 業務設計書"
    ws["B2"].font = title_font

    rows = [
        ("対象ファイル", parse_result.source_file),
        ("生成日時", datetime.now().strftime("%Y年%m月%d日 %H:%M")),
        ("ツールバージョン", "v1.0.0"),
        ("備考", "本設計書はバッチ設計書メーカーにより自動生成されました。\n内容は必ず目視で確認してください。"),
    ]

    for i, (label, value) in enumerate(rows, start=4):
        label_cell = ws.cell(row=i, column=1, value=label)
        label_cell.font = SUBHEADER_FONT
        label_cell.fill = SUBHEADER_FILL
        label_cell.border = thin_border
        label_cell.alignment = Alignment(vertical="center")

        value_cell = ws.cell(row=i, column=2, value=value)
        value_cell.font = BODY_FONT
        value_cell.border = thin_border
        value_cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.row_dimensions[7].height = 40


def _create_io_sheet(wb, parse_result):
    """シート2：I/O定義一覧"""
    ws = wb.create_sheet("IO定義")
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 22

    headers = ["No", "ファイル名（SELECT）", "FD名", "レコード長",
               "物理ファイル名", "ファイル編成", "アクセスモード", "レコードキー"]
    for col, h in enumerate(headers, start=1):
        _apply_header(ws.cell(row=1, column=col), h)

    if not parse_result.file_definitions:
        ws.cell(row=2, column=1).value = "（I/O定義が検出されませんでした）"
        return

    for i, fd in enumerate(parse_result.file_definitions, start=1):
        row = i + 1
        _apply_body(ws.cell(row=row, column=1), i)
        _apply_body(ws.cell(row=row, column=2), fd.file_name or "（不明）")
        _apply_body(ws.cell(row=row, column=3), fd.fd_name or "（不明）")
        _apply_body(ws.cell(row=row, column=4), fd.record_length or "（不明）")
        _apply_body(ws.cell(row=row, column=5), fd.assign_to or "")
        _apply_body(ws.cell(row=row, column=6), fd.organization or "")
        _apply_body(ws.cell(row=row, column=7), fd.access_mode or "")
        _apply_body(ws.cell(row=row, column=8), fd.record_key or "")


def _create_output_sheet(wb, parse_result):
    """シート3：出力ファイル一覧（このプログラムが何を出力するか）"""
    ws = wb.create_sheet("出力ファイル")
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 30

    headers = ["No", "ファイル名", "物理ファイル名", "I/O区分", "ファイル編成", "備考"]
    for col, h in enumerate(headers, start=1):
        _apply_header(ws.cell(row=1, column=col), h)

    output_files = [fd for fd in parse_result.file_definitions
                    if fd.io_mode in ("OUTPUT", "I-O", "EXTEND")]

    if not output_files:
        ws.cell(row=2, column=1).value = "（出力ファイルが検出されませんでした）"
        return

    io_label = {"OUTPUT": "出力", "I-O": "入出力", "EXTEND": "追記", "INPUT": "入力", "": ""}
    for i, fd in enumerate(output_files, start=1):
        row = i + 1
        _apply_body(ws.cell(row=row, column=1), i)
        _apply_body(ws.cell(row=row, column=2), fd.file_name or fd.fd_name or "（不明）")
        _apply_body(ws.cell(row=row, column=3), fd.assign_to or "")
        _apply_body(ws.cell(row=row, column=4), io_label.get(fd.io_mode, fd.io_mode))
        _apply_body(ws.cell(row=row, column=5), fd.organization or "")
        _apply_body(ws.cell(row=row, column=6), "")


FIELD_COLORS = ["D6E4F0", "FFF2CC", "E2EFDA", "FCE4D6", "EAD1DC", "D9EAD3", "CFE2F3", "FFF9C4"]


def _write_layout_cell(ws, row, col_start, col_end, value, fill, font):
    if col_end > col_start:
        ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start)
    cell.value = value
    cell.fill = fill
    cell.font = font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border


def _create_layout_sheets(wb, parse_result):
    """フィールドを持つファイルごとに帳票レイアウトシートを生成"""
    for fd in parse_result.file_definitions:
        if not fd.fields:
            continue
        sheet_name = f"レイアウト_{(fd.file_name or fd.fd_name)[:12]}"
        ws = wb.create_sheet(sheet_name)
        _build_layout_sheet(ws, fd)


def _build_layout_sheet(ws, fd):
    """1ファイル分の帳票レイアウトをExcelシートに描画"""
    file_label = fd.file_name or fd.fd_name
    total_bytes = sum(f.byte_len for f in fd.fields)

    # タイトル行
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(total_bytes, 1))
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"帳票レイアウト：{file_label}　（{fd.assign_to or ''}　レコード長:{total_bytes}バイト）"
    title_cell.font = Font(name="メイリオ", size=11, bold=True, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 20

    # 桁位置マーカー行（5の倍数に番号）
    for col in range(1, total_bytes + 1):
        c = ws.cell(row=2, column=col)
        if col % 5 == 1 or col == 1:
            c.value = col
            c.font = Font(name="メイリオ", size=7, color="888888")
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = 2.2

    ws.row_dimensions[2].height = 14

    # フィールド名行（結合・色分け）
    ws.row_dimensions[3].height = 28
    ws.row_dimensions[4].height = 16
    ws.row_dimensions[5].height = 14

    for i, field in enumerate(fd.fields):
        col_start = field.start_pos
        col_end = field.start_pos + field.byte_len - 1
        color = FIELD_COLORS[i % len(FIELD_COLORS)]
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        _write_layout_cell(ws, 3, col_start, col_end, field.name, fill,
                           Font(name="メイリオ", size=9, bold=True))
        _write_layout_cell(ws, 4, col_start, col_end, f"PIC {field.pic}", fill,
                           Font(name="メイリオ", size=7))
        _write_layout_cell(ws, 5, col_start, col_end, field.byte_len, fill,
                           Font(name="メイリオ", size=7, color="555555"))


def _create_flow_sheet(wb, parse_result):
    """シート3：処理フロー"""
    ws = wb.create_sheet("処理フロー")
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 30

    headers = ["No", "呼び出し元（段落名）", "呼び出し先（段落名）", "種別", "備考"]
    for col, h in enumerate(headers, start=1):
        _apply_header(ws.cell(row=1, column=col), h)

    if not parse_result.perform_entries:
        ws.cell(row=2, column=1).value = "（処理フローが検出されませんでした）"
        return

    for i, entry in enumerate(parse_result.perform_entries, start=1):
        row = i + 1
        _apply_body(ws.cell(row=row, column=1), i)
        _apply_body(ws.cell(row=row, column=2), entry.caller or "（不明）")
        _apply_body(ws.cell(row=row, column=3), entry.callee)
        _apply_body(ws.cell(row=row, column=4), entry.perform_type)
        _apply_body(ws.cell(row=row, column=5), "")


def _create_exception_sheet(wb, parse_result):
    """シート4：例外処理一覧"""
    ws = wb.create_sheet("例外処理")
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 30

    headers = ["No", "発生箇所（段落名）", "例外種別", "処理内容（抜粋）", "備考"]
    for col, h in enumerate(headers, start=1):
        _apply_header(ws.cell(row=1, column=col), h)

    if not parse_result.exception_entries:
        ws.cell(row=2, column=1).value = "（例外処理が検出されませんでした）"
        return

    for i, entry in enumerate(parse_result.exception_entries, start=1):
        row = i + 1
        _apply_body(ws.cell(row=row, column=1), i)
        _apply_body(ws.cell(row=row, column=2), entry.location or "（不明）")
        _apply_body(ws.cell(row=row, column=3), entry.exception_type)
        _apply_body(ws.cell(row=row, column=4), entry.handling)
        _apply_body(ws.cell(row=row, column=5), "")
